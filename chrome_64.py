from selenium import webdriver
from openpyxl import load_workbook


click_me = ["Сервисы для поставщиков и потребителей информации",
            "Новости",
            "Социальный калькулятор",
            "Кабинет поставщика информации",
            "Кабинет органа, назначающего меры социальной поддержки",
            "Личный кабинет гражданина",
            "Кабинет аналитика"]


def ft_load_time(click_me):
    driver = webdriver.Chrome()
    driver.get('http://egisso.ru/site/')
    test_page = driver.find_element_by_link_text(click_me)
    # driver.refresh()
    test_page.click()
    # time.sleep(5) # - не влияет на load time
    load_time = driver.execute_script(
        "return (window.performance.timing.loadEventEnd - window.performance.timing.navigationStart);")
    driver.close()
    return str(load_time)


res = []
for link in click_me:
    res.append(ft_load_time(link))

wb = load_workbook('info.XLSX')
#wb.create_sheet(title='Время отклика', index=0)
sheet = wb['Время отклика']
sheet['C3'] = res[0]
sheet['C11'] = res[1]
sheet['C19'] = res[2]
sheet['C27'] = res[3]
sheet['C35'] = res[4]
sheet['C43'] = res[5]
sheet['C51'] = res[6]
wb.save('info.XLSX')