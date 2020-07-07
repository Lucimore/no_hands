from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import WebDriverException

click_me = ["Сервисы для поставщиков и потребителей информации",
            "Новости",
            "Социальный калькулятор",
            "Кабинет поставщика информации",
            "Кабинет органа, назначающего меры социальной поддержки",
            "Личный кабинет гражданина",
            "Кабинет аналитика"]


def ft_load_time(click_me):
    driver = webdriver.Firefox()
    driver.get('http://egisso.ru/site/')
    test_page = driver.find_element_by_link_text(click_me)
    # driver.refresh()
    try:
        test_page.click()
        driver.refresh()
        # time.sleep(5)  # - не влияет на load time
        load_time = driver.execute_script(
            "return (window.performance.timing.loadEventEnd - window.performance.timing.navigationStart);")
        driver.close()
        return str(load_time)
    except WebDriverException:
        driver.close()
        return "There was an ERROR"


res = []
for link in click_me:
    res.append(ft_load_time(link))

wb = load_workbook('info.XLSX')
# wb.create_sheet(title='Время отклика', index=0)
sheet = wb['Время отклика']
sheet['F2'] = res[0]
sheet['F10'] = res[1]
sheet['F18'] = res[2]
sheet['F26'] = res[3]
sheet['F34'] = res[4]
sheet['F42'] = res[5]
sheet['F50'] = res[6]
wb.save('info.XLSX')
