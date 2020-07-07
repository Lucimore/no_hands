import time
from selenium import webdriver
from openpyxl import load_workbook

click_me = ["Сервисы для поставщиков и потребителей информации",
            "Новости",
            "Социальный калькулятор",
            "Кабинет поставщика информации",
            "Кабинет органа, назначающего меры социальной поддержки",
            "Личный кабинет гражданина",
            "Кабинет аналитика"]


def ft_load_time(click_me):
    driver = webdriver.Edge("msedgedriver32.exe")
    driver.get('http://egisso.ru/site/')
    test_page = driver.find_element_by_link_text(click_me)
    # driver.refresh()
    test_page.click()
    time.sleep(5)
    load_time = driver.execute_script(
        "return (window.performance.timing.loadEventEnd - window.performance.timing.navigationStart);")
    driver.close()
    return str(load_time)


res = []
for link in click_me:
    res.append(ft_load_time(link))

wb = load_workbook('info.XLSX')
# wb.create_sheet(title='Время отклика', index=0)
sheet = wb['Время отклика']
sheet['E2'] = res[0]
sheet['E10'] = res[1]
sheet['E18'] = res[2]
sheet['E26'] = res[3]
sheet['E34'] = res[4]
sheet['E42'] = res[5]
sheet['E50'] = res[6]
wb.save('info.XLSX')
